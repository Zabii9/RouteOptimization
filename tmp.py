"""
Route Optimizer — OSRM Nearest Neighbor
+ Interactive HTML Map (Leaflet + OSRM road geometry, all pins)
+ Excel report with Google Maps links

Run:
    pip install requests pandas openpyxl
    python route_optimizer_osrm.py

Output:
    optimized_route_osrm.xlsx    — full report + Maps links
    route_map_osrm.html          — interactive map, open in any browser
"""

import requests
import pandas as pd
import json
import webbrowser
import os
import time
from datetime import datetime

# ──────────────────────────────────────────────────────────────
# CONFIGURATION
# ──────────────────────────────────────────────────────────────
OSRM_BASE_URL = "http://router.project-osrm.org"
# Local server?  → OSRM_BASE_URL = "http://localhost:5000"

WAREHOUSE = {
    "name": "Warehouse (Start/End)",
    "lat":  24.8542977,
    "lon":  67.1534304,
}

# ──────────────────────────────────────────────────────────────
# SHOP DATA  (paste your Load Form rows here)
# ──────────────────────────────────────────────────────────────
shops_data = [
    {"ShopCode": "N00000015269", "ShopName": "MADINA STOR",         "Lat": 24.9028183, "Lon": 67.2066233, "OrderValue": 100.61},
    {"ShopCode": "N00000120011", "ShopName": "MS CARNAL STORE",     "Lat": 24.8998067, "Lon": 67.2058167, "OrderValue": 318.11},
    {"ShopCode": "N00000120024", "ShopName": "KHAN STORE",          "Lat": 24.9024085, "Lon": 67.2089997, "OrderValue": 100.61},
    {"ShopCode": "N00000120030", "ShopName": "AL AKBAR BAKRE",      "Lat": 24.9016217, "Lon": 67.21019,   "OrderValue": 469.72},
    {"ShopCode": "N00000125205", "ShopName": "ALI COLD",            "Lat": 24.9018283, "Lon": 67.20937,   "OrderValue": 460.16},
    {"ShopCode": "N00000133940", "ShopName": "Maa g stor",          "Lat": 24.9022267, "Lon": 67.2063233, "OrderValue": 100.61},
    {"ShopCode": "N00000133941", "ShopName": "FRIEND POINT",        "Lat": 24.9010004, "Lon": 67.2088899, "OrderValue": 110.17},
    {"ShopCode": "N00000133943", "ShopName": "ABDUL MAJEED",        "Lat": 24.8996533, "Lon": 67.20562,   "OrderValue": 671.83},
    {"ShopCode": "N00000134024", "ShopName": "Rashid Minhas store", "Lat": 24.9018585, "Lon": 67.2080132, "OrderValue": 1694.29},
    {"ShopCode": "N00000134025", "ShopName": "Al Ameen medical",    "Lat": 24.9026633, "Lon": 67.2067517, "OrderValue": 10439.55},
    {"ShopCode": "N00000134031", "ShopName": "ATTARI G STORE",      "Lat": 24.9018276, "Lon": 67.2094566, "OrderValue": 100.61},
    {"ShopCode": "N00000134032", "ShopName": "Variety store",       "Lat": 24.90185,   "Lon": 67.20941,   "OrderValue": 91.05},
    {"ShopCode": "N00000134038", "ShopName": "ARSALAN STORE",       "Lat": 24.9005247, "Lon": 67.2101616, "OrderValue": 100.61},
    {"ShopCode": "N00000149854", "ShopName": "S.N.S G STOR",        "Lat": 24.9008183, "Lon": 67.2036667, "OrderValue": 9.56},
    {"ShopCode": "N00000149859", "ShopName": "Super GS",            "Lat": 24.9015199, "Lon": 67.2102531, "OrderValue": 9.56},
    {"ShopCode": "N00000184206", "ShopName": "WAA STORE",           "Lat": 24.9007797, "Lon": 67.2045995, "OrderValue": 9.56},
    {"ShopCode": "N00000184818", "ShopName": "IKRAM MILK",          "Lat": 24.8995324, "Lon": 67.204844,  "OrderValue": 318.11},
    {"ShopCode": "N00000186252", "ShopName": "BABA FAREED",         "Lat": 24.9013724, "Lon": 67.2032915, "OrderValue": 318.11},
    {"ShopCode": "N00000312632", "ShopName": "RAZA STOR",           "Lat": 24.900865,  "Lon": 67.2035665, "OrderValue": 100.61},
    {"ShopCode": "N00000328463", "ShopName": "MADINA STOR (2)",     "Lat": 24.9026633, "Lon": 67.2067517, "OrderValue": 100.61},
    {"ShopCode": "N00000328464", "ShopName": "Hashmi Medical",      "Lat": 24.9024283, "Lon": 67.2074133, "OrderValue": 9581.61},
    {"ShopCode": "N00000328465", "ShopName": "Hasan super store",   "Lat": 24.9025683, "Lon": 67.206735,  "OrderValue": 100.61},
    {"ShopCode": "N00000328468", "ShopName": "Atari Store 2",       "Lat": 24.9030183, "Lon": 67.2062283, "OrderValue": 100.61},
    {"ShopCode": "N00000328473", "ShopName": "Anmol bakery",        "Lat": 24.8972733, "Lon": 67.2108383, "OrderValue": 318.11},
    {"ShopCode": "N00000328474", "ShopName": "A J STOR",            "Lat": 24.900765,  "Lon": 67.209045,  "OrderValue": 3930.27},
    {"ShopCode": "N00000328475", "ShopName": "Kareem store",        "Lat": 24.9022733, "Lon": 67.2075267, "OrderValue": 100.61},
    {"ShopCode": "N00000330287", "ShopName": "WASEEM",              "Lat": 24.8974483, "Lon": 67.2076817, "OrderValue": 9.56},
    {"ShopCode": "N00000347421", "ShopName": "Fahad store",         "Lat": 24.8970367, "Lon": 67.2078983, "OrderValue": 318.11},
    {"ShopCode": "N00000347422", "ShopName": "Bilal Store",         "Lat": 24.8969159, "Lon": 67.2080564, "OrderValue": 318.11},
    {"ShopCode": "N00000347423", "ShopName": "Sabor Store",         "Lat": 24.898445,  "Lon": 67.2072017, "OrderValue": 9.56},
    {"ShopCode": "N00000347424", "ShopName": "NADEEM STORE",        "Lat": 24.8977567, "Lon": 67.210555,  "OrderValue": 1417.59},
    {"ShopCode": "N00000330269", "ShopName": "Al madina pan shop",  "Lat": 24.9008017, "Lon": 67.2157867, "OrderValue": 318.11},
    {"ShopCode": "N00000002972", "ShopName": "D SHAHAB GS",         "Lat": 24.9022617, "Lon": 67.2076067, "OrderValue": 100.61},
    {"ShopCode": "N00000003788", "ShopName": "C IKRAM SWEET",       "Lat": 24.896655,  "Lon": 67.2092683, "OrderValue": 5128.24},
    {"ShopCode": "N00000003858", "ShopName": "C SELACTION BAKERY",  "Lat": 24.9007067, "Lon": 67.204675,  "OrderValue": 318.11},
    {"ShopCode": "N00000003880", "ShopName": "IQRA STORE",          "Lat": 24.8967199, "Lon": 67.2085883, "OrderValue": 2536.59},
    {"ShopCode": "N00000003900", "ShopName": "A JWA SWEET",         "Lat": 24.9002117, "Lon": 67.2097917, "OrderValue": 4936.67},
    {"ShopCode": "N00000003915", "ShopName": "C FAISAL MEDICAL",    "Lat": 24.9025783, "Lon": 67.207075,  "OrderValue": 2784.72},
    {"ShopCode": "N00000003918", "ShopName": "B IMRAN STORE",       "Lat": 24.8982633, "Lon": 67.2104933, "OrderValue": 318.11},
    {"ShopCode": "N00000003923", "ShopName": "D AL QAMMR",          "Lat": 24.89828,   "Lon": 67.21041,   "OrderValue": 318.11},
    {"ShopCode": "N00000003930", "ShopName": "C TAWAKKAL 2",        "Lat": 24.9006919, "Lon": 67.2057848, "OrderValue": 318.11},
    {"ShopCode": "N00000007436", "ShopName": "C ZEESHAN GST",       "Lat": 24.9013883, "Lon": 67.2085883, "OrderValue": 100.61},
    {"ShopCode": "N00000013330", "ShopName": "husani milak",        "Lat": 24.9022189, "Lon": 67.2064691, "OrderValue": 2259.74},
    {"ShopCode": "N00000013345", "ShopName": "C WAHAB STORE",       "Lat": 24.9008067, "Lon": 67.203725,  "OrderValue": 318.11},
    {"ShopCode": "N00000120768", "ShopName": "JUNAID STOR",         "Lat": 24.9002417, "Lon": 67.2149867, "OrderValue": 318.11},
]


# ──────────────────────────────────────────────────────────────
# OSRM — DISTANCE MATRIX  (one bulk Table API call)
# ──────────────────────────────────────────────────────────────

def get_osrm_table(coords, retries=3):
    """
    coords : list of (lat, lon)
    Returns (dist_km_matrix, dur_min_matrix) or (None, None)
    """
    coord_str = ";".join(f"{lon},{lat}" for lat, lon in coords)
    url = f"{OSRM_BASE_URL}/table/v1/driving/{coord_str}?annotations=distance,duration"
    for attempt in range(retries):
        try:
            resp = requests.get(url, timeout=90)
            resp.raise_for_status()
            data = resp.json()
            if data.get("code") == "Ok":
                dist = [[v / 1000.0 if v is not None else 1e9 for v in row]
                        for row in data["distances"]]
                dur  = [[v / 60.0  if v is not None else 1e9 for v in row]
                        for row in data["durations"]]
                return dist, dur
        except Exception as e:
            print(f"  [OSRM Table] attempt {attempt+1} failed: {e}")
            time.sleep(2)
    return None, None


# ──────────────────────────────────────────────────────────────
# OSRM — FULL ROAD GEOMETRY  (actual polyline for the map)
# ──────────────────────────────────────────────────────────────

def get_osrm_route_geometry(ordered_stops, retries=3):
    """
    ordered_stops : list of dicts with Lat/Lon or lat/lon keys
    Returns list of [lat, lon] pairs (road geometry) or []
    """
    def _lat(s): return s.get("lat", s.get("Lat"))
    def _lon(s): return s.get("lon", s.get("Lon"))

    coord_str = ";".join(f"{_lon(s)},{_lat(s)}" for s in ordered_stops)
    url = (f"{OSRM_BASE_URL}/route/v1/driving/{coord_str}"
           f"?overview=full&geometries=geojson&steps=false")
    for attempt in range(retries):
        try:
            resp = requests.get(url, timeout=90)
            resp.raise_for_status()
            data = resp.json()
            if data.get("code") == "Ok":
                coords = data["routes"][0]["geometry"]["coordinates"]
                return [[c[1], c[0]] for c in coords]   # GeoJSON lon,lat → lat,lon
        except Exception as e:
            print(f"  [OSRM Geometry] attempt {attempt+1} failed: {e}")
            time.sleep(2)
    return []


# ──────────────────────────────────────────────────────────────
# NEAREST NEIGHBOR ROUTE OPTIMIZER
# ──────────────────────────────────────────────────────────────

def haversine_fallback(coords):
    """Build a straight-line distance matrix when OSRM is unavailable."""
    import math
    n = len(coords)
    def hav(a, b):
        R = 6371
        dlat = math.radians(b[0] - a[0]); dlon = math.radians(b[1] - a[1])
        x = (math.sin(dlat/2)**2
             + math.cos(math.radians(a[0])) * math.cos(math.radians(b[0]))
             * math.sin(dlon/2)**2)
        return R * 2 * math.asin(math.sqrt(x))
    dist = [[hav(coords[i], coords[j]) for j in range(n)] for i in range(n)]
    dur  = [[dist[i][j] / 40 * 60    for j in range(n)] for i in range(n)]
    return dist, dur


def optimize_route(shops, warehouse):
    print("\n" + "=" * 65)
    print("  ROUTE OPTIMIZER — OSRM Nearest Neighbor")
    print("=" * 65)
    print(f"  Warehouse : {warehouse['name']}")
    print(f"  Shops     : {len(shops)}")
    print(f"  Time      : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 65)

    all_coords = [(warehouse["lat"], warehouse["lon"])] + \
                 [(s["Lat"], s["Lon"]) for s in shops]
    n = len(all_coords)

    print(f"\n[1/4] Fetching {n}x{n} distance matrix from OSRM ...")
    dist_matrix, dur_matrix = get_osrm_table(all_coords)

    if dist_matrix is None:
        print("  [!] OSRM unavailable — using Haversine straight-line fallback.")
        dist_matrix, dur_matrix = haversine_fallback(all_coords)
    else:
        print("  [OK] Distance matrix ready.")

    # ── Nearest Neighbour ────────────────────────────────────
    print("\n[2/4] Running Nearest Neighbor ...")
    visited   = [False] * n
    route_idx = [0]; visited[0] = True
    unvisited = list(range(1, n))
    while unvisited:
        cur     = route_idx[-1]
        nearest = min(unvisited, key=lambda j: dist_matrix[cur][j])
        route_idx.append(nearest)
        visited[nearest] = True
        unvisited.remove(nearest)
    route_idx.append(0)
    print("  [OK] Route optimized.")

    # ── Build result rows ────────────────────────────────────
    rows = []
    cum_km = cum_min = 0.0
    for step, (f, t) in enumerate(zip(route_idx[:-1], route_idx[1:])):
        leg_km  = dist_matrix[f][t]
        leg_min = dur_matrix[f][t]
        cum_km  += leg_km
        cum_min += leg_min
        from_name = warehouse["name"] if f == 0 else shops[f - 1]["ShopName"]
        to_name   = warehouse["name"] if t == 0 else shops[t - 1]["ShopName"]
        to_code   = "WAREHOUSE"        if t == 0 else shops[t - 1]["ShopCode"]
        to_val    = ""                 if t == 0 else shops[t - 1]["OrderValue"]
        rows.append({
            "Stop #":          step + 1,
            "From":            from_name,
            "To (Shop)":       to_name,
            "ShopCode":        to_code,
            "OrderValue (Rs)": to_val,
            "Leg km":          round(leg_km,  3),
            "Leg min":         round(leg_min, 1),
            "Cumulative km":   round(cum_km,  3),
            "Cumulative min":  round(cum_min, 1),
        })

    return rows, cum_km, cum_min, route_idx


def print_route_table(rows, total_km, total_min):
    df = pd.DataFrame(rows)
    pd.set_option("display.max_rows", 200)
    pd.set_option("display.width",    220)
    print("\n" + "=" * 65)
    print("  OPTIMIZED DELIVERY ROUTE")
    print("=" * 65)
    print(df.to_string(index=False))
    print("=" * 65)
    print(f"  TOTAL ROAD DISTANCE : {total_km:.3f} km")
    print(f"  TOTAL DRIVE TIME    : {total_min:.1f} min  ({total_min/60:.1f} hrs)")
    print("=" * 65)


# ──────────────────────────────────────────────────────────────
# GOOGLE MAPS LINKS  (chunked, max 9 stops per link)
# ──────────────────────────────────────────────────────────────

def build_gmaps_chunks(ordered_shops, warehouse, chunk_size=9):
    def _lat(s): return s.get("lat", s.get("Lat"))
    def _lon(s): return s.get("lon", s.get("Lon"))
    def _name(s): return s.get("ShopName", s.get("name", "Warehouse"))

    all_stops = [warehouse] + ordered_shops + [warehouse]
    chunks, i, num = [], 0, 1
    while i < len(all_stops) - 1:
        end  = min(i + chunk_size + 1, len(all_stops) - 1)
        seg  = all_stops[i:end + 1]
        orig = f"{_lat(seg[0])},{_lon(seg[0])}"
        dest = f"{_lat(seg[-1])},{_lon(seg[-1])}"
        wps  = "|".join(f"{_lat(s)},{_lon(s)}" for s in seg[1:-1])
        url  = (f"https://www.google.com/maps/dir/?api=1"
                f"&origin={orig}&destination={dest}&travelmode=driving"
                + (f"&waypoints={wps}" if wps else ""))
        chunks.append({"num": num, "from": _name(seg[0]), "to": _name(seg[-1]),
                       "stops": f"Stops {i}-{end}", "url": url, "segment": seg})
        i = end; num += 1
    return chunks


# ──────────────────────────────────────────────────────────────
# HTML MAP  — Leaflet + OSRM road geometry  (self-contained)
# ──────────────────────────────────────────────────────────────

def generate_html_map(ordered_shops, warehouse, road_geometry,
                      total_km, total_min,
                      load_form="D0573LF11548",
                      deliveryman="Mohid DM [D0573OB45]"):

    # Build stop list for JS
    stops_js = [{"idx": 0, "label": "W", "name": warehouse["name"],
                  "code": "WAREHOUSE", "lat": warehouse["lat"],
                  "lon": warehouse["lon"], "val": 0, "type": "warehouse", "stop": 0}]
    for i, s in enumerate(ordered_shops, 1):
        stops_js.append({
            "idx": i, "label": str(i),
            "name": s["ShopName"], "code": s["ShopCode"],
            "lat": s["Lat"], "lon": s["Lon"], "val": s["OrderValue"],
            "type": "high" if s["OrderValue"] > 1000 else "shop", "stop": i,
        })

    total_order = sum(s["OrderValue"] for s in ordered_shops)

    if road_geometry:
        road_json   = json.dumps(road_geometry)
        road_note   = "OSRM actual road geometry"
    else:
        straight = ([[warehouse["lat"], warehouse["lon"]]]
                    + [[s["Lat"], s["Lon"]] for s in ordered_shops]
                    + [[warehouse["lat"], warehouse["lon"]]])
        road_json = json.dumps(straight)
        road_note = "straight-line fallback"

    gen_time = datetime.now().strftime("%d %b %Y  %H:%M")

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Route Map — {load_form}</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;
      display:flex;height:100vh;overflow:hidden;background:#f4f4f4}}

#sidebar{{width:310px;min-width:260px;height:100vh;display:flex;
          flex-direction:column;background:#fff;
          border-right:1px solid #ddd;z-index:1000;flex-shrink:0}}

#hdr{{padding:14px 16px;background:#1a1a2e;color:#fff;flex-shrink:0}}
#hdr h2{{font-size:14px;font-weight:600;margin-bottom:3px}}
#hdr p{{font-size:11px;opacity:.65;line-height:1.6}}

.stats{{display:grid;grid-template-columns:repeat(3,1fr);gap:0;
        border-bottom:1px solid #eee;flex-shrink:0}}
.stat{{padding:10px 0;text-align:center;border-right:1px solid #eee}}
.stat:last-child{{border-right:none}}
.sv{{font-size:17px;font-weight:600;color:#1a1a2e}}
.sl{{font-size:10px;color:#999;margin-top:1px}}

#list{{overflow-y:auto;flex:1}}
.row{{display:flex;align-items:center;gap:10px;padding:9px 14px;
      cursor:pointer;border-bottom:1px solid #f2f2f2;transition:background .12s}}
.row:hover{{background:#f0f4ff}}
.row.active{{background:#e8eeff}}
.pin{{width:28px;height:28px;border-radius:50%;display:flex;align-items:center;
       justify-content:center;font-size:10px;font-weight:700;color:#fff;flex-shrink:0}}
.wh{{background:#1D9E75}}.sh{{background:#378ADD}}.hi{{background:#E24B4A}}
.info{{flex:1;min-width:0}}
.sname{{font-size:12px;font-weight:500;color:#111;
        white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
.scode{{font-size:10px;color:#aaa;margin-top:1px}}
.sval{{font-size:11px;font-weight:600;color:#1a1a2e;white-space:nowrap}}

#map{{flex:1;height:100vh}}

.legend{{background:rgba(255,255,255,.93);padding:9px 12px;
         border-radius:7px;box-shadow:0 1px 5px rgba(0,0,0,.18);font-size:11px;line-height:2}}
.lr{{display:flex;align-items:center;gap:7px}}
.ld{{width:10px;height:10px;border-radius:50%}}
.ll{{width:18px;height:3px;border-radius:2px}}

.note-bar{{font-size:10px;color:#aaa;padding:5px 14px;
           border-top:1px solid #eee;flex-shrink:0;text-align:center}}
</style>
</head>
<body>

<div id="sidebar">
  <div id="hdr">
    <h2>&#x1F4E6; {load_form}</h2>
    <p>{deliveryman}<br>{gen_time}</p>
  </div>
  <div class="stats">
    <div class="stat"><div class="sv">{len(ordered_shops)}</div><div class="sl">Shops</div></div>
    <div class="stat"><div class="sv">{total_km:.1f}</div><div class="sl">km</div></div>
    <div class="stat"><div class="sv">{total_min:.0f}</div><div class="sl">min</div></div>
  </div>
  <div id="list"></div>
  <div class="note-bar">Route: {road_note}</div>
</div>

<div id="map"></div>

<script>
const STOPS={json.dumps(stops_js, ensure_ascii=False)};
const ROAD={road_json};

const map=L.map('map').setView([24.895,67.207],14);
L.tileLayer('https://{{s}}.tile.openstreetmap.org/{{z}}/{{x}}/{{y}}.png',{{
  attribution:'&copy; <a href="https://www.openstreetmap.org/copyright">OpenStreetMap</a> contributors',
  maxZoom:19
}}).addTo(map);

// Road route line
if(ROAD.length>1){{
  L.polyline(ROAD,{{color:'#534AB7',weight:4,opacity:.85,dashArray:'9,4'}}).addTo(map);
}}

// Markers
function mkIcon(label,type){{
  const c={{warehouse:'#1D9E75',shop:'#378ADD',high:'#E24B4A'}}[type]||'#888';
  const fs=label.length>2?7:label.length>1?9:11;
  return L.divIcon({{
    className:'',
    html:`<div style="width:30px;height:30px;border-radius:50%;background:${{c}};
      border:2.5px solid #fff;display:flex;align-items:center;justify-content:center;
      font-size:${{fs}}px;font-weight:700;color:#fff;
      box-shadow:0 2px 6px rgba(0,0,0,.3)">${{label}}</div>`,
    iconSize:[30,30],iconAnchor:[15,15],popupAnchor:[0,-16]
  }});
}}

const MRK={{}};
STOPS.forEach(s=>{{
  const m=L.marker([s.lat,s.lon],{{icon:mkIcon(s.label,s.type)}}).addTo(map);
  const vl=s.val>0?`<br><b>Order:</b> Rs ${{s.val.toLocaleString()}}`:'';
  m.bindPopup(`<div style="font-size:13px;min-width:170px">
    <b style="font-size:14px">${{s.name}}</b><br>
    <span style="color:#999;font-size:11px">${{s.code}}</span>
    ${{vl}}<br>
    <span style="font-size:10px;color:#bbb">${{s.lat.toFixed(6)}}, ${{s.lon.toFixed(6)}}</span>
  </div>`);
  MRK[s.idx]=m;
}});

// Sidebar list
const listEl=document.getElementById('list');
STOPS.forEach(s=>{{
  const r=document.createElement('div');
  r.className='row'; r.dataset.idx=s.idx;
  const pc={{warehouse:'wh',shop:'sh',high:'hi'}}[s.type]||'sh';
  r.innerHTML=`
    <div class="pin ${{pc}}">${{s.label}}</div>
    <div class="info">
      <div class="sname">${{s.name}}</div>
      <div class="scode">${{s.code}}</div>
    </div>
    ${{s.val>0?`<div class="sval">Rs ${{s.val.toLocaleString()}}</div>`:''}}`;
  r.addEventListener('click',()=>{{
    document.querySelectorAll('.row').forEach(x=>x.classList.remove('active'));
    r.classList.add('active');
    map.flyTo([s.lat,s.lon],17,{{duration:1}});
    MRK[s.idx].openPopup();
  }});
  listEl.appendChild(r);
}});

// Legend
const leg=L.control({{position:'bottomright'}});
leg.onAdd=()=>{{
  const d=L.DomUtil.create('div','legend');
  d.innerHTML=`
    <div class="lr"><div class="ld" style="background:#1D9E75"></div>Warehouse</div>
    <div class="lr"><div class="ld" style="background:#378ADD"></div>Shop (normal)</div>
    <div class="lr"><div class="ld" style="background:#E24B4A"></div>High value (&gt;Rs 1000)</div>
    <div class="lr"><div class="ll" style="background:#534AB7"></div>OSRM road route</div>`;
  return d;
}};
leg.addTo(map);
</script>
</body>
</html>"""


# ──────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ──────────────────────────────────────────────────────────────

def export_excel(rows, total_km, total_min, maps_chunks,
                 filename="optimized_route_osrm.xlsx"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb  = Workbook()
    hfill = PatternFill("solid", fgColor="1A1A2E")

    # ── Sheet 1: Route ──────────────────────────────────────
    ws = wb.active
    ws.title = "Optimized Route"
    cols = list(rows[0].keys())
    for ci, c in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")

    for ri, row in enumerate(rows, 2):
        for ci, c in enumerate(cols, 1):
            ws.cell(row=ri, column=ci, value=row[c])

    sr = len(rows) + 2
    ws.cell(row=sr, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=sr, column=6, value=round(total_km,  3))
    ws.cell(row=sr, column=7, value=round(total_min, 1))

    for col in ws.columns:
        w = max((len(str(c.value)) for c in col if c.value), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(w + 4, 42)

    # ── Sheet 2: Maps Links ─────────────────────────────────
    ws2 = wb.create_sheet("Google Maps Links")
    for ci, h in enumerate(["Link #", "Stops", "From", "To", "Open in Maps"], 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hfill

    for ri, ch in enumerate(maps_chunks, 2):
        ws2.cell(row=ri, column=1, value=f"Link {ch['num']}")
        ws2.cell(row=ri, column=2, value=ch["stops"])
        ws2.cell(row=ri, column=3, value=ch["from"])
        ws2.cell(row=ri, column=4, value=ch["to"])
        lc = ws2.cell(row=ri, column=5, value=f"Open Link {ch['num']} in Maps")
        lc.hyperlink = ch["url"]
        lc.font      = Font(color="1155CC", underline="single")

    for ci, w in enumerate([10, 14, 32, 32, 30], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    wb.save(filename)
    print(f"  [OK] Excel saved  ->  {filename}")


# ──────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────

if __name__ == "__main__":
    # Ensure dependencies
    try:
        import openpyxl
    except ImportError:
        import subprocess, sys
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])

    # 1. Optimize route via OSRM distance matrix
    route_rows, total_km, total_min, route_idx = optimize_route(shops_data, WAREHOUSE)
    print_route_table(route_rows, total_km, total_min)

    # 2. Rebuild ordered shop list from optimized index
    #    route_idx = [0, s1, s2, ..., sN, 0]  (0 = warehouse index)
    ordered_shops = [shops_data[i - 1] for i in route_idx[1:-1]]

    # 3. Fetch OSRM road geometry (actual road polyline for map)
    print("\n[3/4] Fetching OSRM road geometry for map ...")
    full_route_stops = [WAREHOUSE] + ordered_shops + [WAREHOUSE]
    road_geom = get_osrm_route_geometry(full_route_stops)
    if road_geom:
        print(f"  [OK] Road geometry: {len(road_geom)} coordinate points.")
    else:
        print("  [!] Road geometry unavailable — straight lines will be drawn on map.")

    # 4. Generate HTML map and open in browser
    print("\n[4/4] Generating interactive HTML map ...")
    html = generate_html_map(
        ordered_shops = ordered_shops,
        warehouse     = WAREHOUSE,
        road_geometry = road_geom,
        total_km      = total_km,
        total_min     = total_min,
    )
    map_file = "route_map_osrm.html"
    with open(map_file, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  [OK] HTML map saved  ->  {map_file}")

    # 5. Build Google Maps chunks and export Excel
    maps_chunks = build_gmaps_chunks(ordered_shops, WAREHOUSE, chunk_size=9)
    export_excel(route_rows, total_km, total_min, maps_chunks)

    # 6. Auto-open map in default browser
    abs_path = os.path.abspath(map_file)
    webbrowser.open(f"file://{abs_path}")

    print("\n" + "=" * 65)
    print("  ALL DONE")
    print(f"  route_map_osrm.html          <- open in any browser")
    print(f"  optimized_route_osrm.xlsx    <- report + clickable Maps links")
    print("=" * 65)